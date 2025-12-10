# sharepoint_helper.py
from asyncio import tasks
import os
import requests
from msal import ConfidentialClientApplication
from dotenv import load_dotenv
import os
import requests
import pandas as pd
from datetime import datetime
import pytz
from collections import defaultdict
import re
import base64
import json
import openpyxl
from openpyxl.utils import get_column_letter
import openai
# ----------------------------
# Load environment variables
# ----------------------------
load_dotenv(override=True)

CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
TENANT_ID = os.getenv("TENANT_ID")

if not all([CLIENT_ID, CLIENT_SECRET, TENANT_ID]):
    raise ValueError("CLIENT_ID, CLIENT_SECRET, and TENANT_ID must be set in .env")

GRAPH_API = "https://graph.microsoft.com/v1.0"

# --- Configuration for OneDrive Feature ---

CLIENT_ID_ONEDRIVE = os.getenv("CLIENT_ID")
CLIENT_SECRET_ONEDRIVE = os.getenv("CLIENT_SECRET")
TENANT_ID_ONEDRIVE = os.getenv("TENANT_ID")
ONEDRIVE_PRIMARY_USER_ID = os.getenv("ONEDRIVE_PRIMARY_USER_ID")
AUTHORITY_ONEDRIVE = f"https://login.microsoftonline.com/{TENANT_ID_ONEDRIVE}"
SCOPE_ONEDRIVE = ["https://graph.microsoft.com/.default"]
ONEDRIVE_USER_ID = os.getenv("ONEDRIVE_USER_ID")
FILE_PATH = os.getenv("ONEDRIVE_FILE_PATH", "Contacts.xlsx")
WORKSHEET_NAME = os.getenv("ONEDRIVE_WORKSHEET_NAME", "Sheet1")
CONTACT_WORKSHEET_NAME= "Sheet1"
GRAPH_API_ENDPOINT = "https://graph.microsoft.com/v1.0"

onedrive_msal_app = ConfidentialClientApplication(
    CLIENT_ID_ONEDRIVE,
    authority=AUTHORITY_ONEDRIVE,
    client_credential=CLIENT_SECRET_ONEDRIVE
)

# ----------------------------
# MSAL Authentication
# ----------------------------
def get_access_token():
    """Returns a Graph API access token using client credentials."""
    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    app = ConfidentialClientApplication(
        CLIENT_ID,
        authority=authority,
        client_credential=CLIENT_SECRET
    )
    token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in token_response:
        raise Exception(f"Failed to get access token: {token_response}")
    return token_response["access_token"]

# ----------------------------
# Fetch all users in org
# ----------------------------
def get_all_users(access_token):
    """Fetch all users and return a cache: user_id -> displayName"""
    headers = {"Authorization": f"Bearer {access_token}"}
    users = []
    url = f"{GRAPH_API}/users?$select=id,displayName,mail"
    while url:
        resp = requests.get(url, headers=headers)
        if resp.status_code != 200:
            raise Exception(f"Error fetching users: {resp.text}")
        data = resp.json()
        users.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return {u['id']: u.get('displayName', u.get('mail')) for u in users}


# ----------------------------
# Get SharePoint site ID
# ----------------------------
def get_site_id(access_token, site_domain, site_path):
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"{GRAPH_API}/sites/{site_domain}:{site_path}"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        raise Exception(f"Error fetching site ID: {resp.text}")
    return resp.json().get("id")


# ----------------------------
# Get SharePoint list ID
# ----------------------------
def get_list_id(access_token, site_id, list_name):
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"{GRAPH_API}/sites/{site_id}/lists"
    resp = requests.get(url, headers=headers)
    if resp.status_code != 200:
        raise Exception(f"Error fetching lists: {resp.text}")
    for l in resp.json().get("value", []):
        if l.get("name") == list_name:
            return l.get("id")
    raise Exception(f"List '{list_name}' not found.")


# ----------------------------
# Get list items
# ----------------------------
def get_list_items(access_token, site_id, list_id):
    headers = {"Authorization": f"Bearer {access_token}"}
    items = []
    url = f"{GRAPH_API}/sites/{site_id}/lists/{list_id}/items?expand=fields($expand=AssignedTo,Author,Editor)"
    while url:
        resp = requests.get(url, headers=headers)
        if resp.status_code != 200:
            raise Exception(f"Error fetching items: {resp.text}")
        data = resp.json()
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


# ----------------------------
# Flatten fields
# ----------------------------
def flatten_fields(item_fields, user_cache=None):
    flat = {}
    for k, v in item_fields.items():
        if isinstance(v, dict):
            lookup_id = v.get("id") or v.get("userId")
            if lookup_id and user_cache:
                flat[k] = user_cache.get(lookup_id, v.get("displayName") or str(v))
            else:
                flat[k] = v.get("lookupValue") or v.get("displayName") or str(v)
        elif isinstance(v, list):
            names = []
            for i in v:
                if isinstance(i, dict):
                    lookup_id = i.get("id") or i.get("userId")
                    if lookup_id and user_cache:
                        names.append(user_cache.get(lookup_id, i.get("displayName") or str(i)))
                    else:
                        names.append(i.get("lookupValue") or i.get("displayName") or str(i))
                else:
                    names.append(str(i))
            flat[k] = ", ".join(names)
        else:
            flat[k] = v
    return flat


# ----------------------------
# Fetch SharePoint list items as structured data
# ----------------------------
def fetch_sharepoint_list(site_domain, site_path, list_name):
    """
    Returns a list of flattened SharePoint list items with user display names.
    
    Example usage:
        items = fetch_sharepoint_list("hamdaz1.sharepoint.com", "/sites/ProposalTeam", "Proposals")
    """
    access_token = get_access_token()
    user_cache = get_all_users(access_token)

    site_id = get_site_id(access_token, site_domain, site_path)
    list_id = get_list_id(access_token, site_id, list_name)
    raw_items = get_list_items(access_token, site_id, list_id)

    structured_items = [flatten_fields(item.get("fields", {}), user_cache) for item in raw_items]
    return structured_items



def items_to_dataframe(items):
    """
    Convert a list of SharePoint item dictionaries into a Pandas DataFrame.
    
    Args:
        items (list): List of dictionaries, each representing a SharePoint item.
        
    Returns:
        pd.DataFrame: DataFrame with all fields as columns.
    """
    if not items:
        return pd.DataFrame()  # Return empty DF if no items
    
    df = pd.DataFrame(items)
    
    # Convert common date columns to datetime, if they exist
    date_cols = ['DueDate', 'BCD', 'Created', 'Modified']
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    
    return df


def compute_overall_analytics(df, period=None):
    """
    Compute overall analytics with optional period filtering.
    
    Args:
        df (pd.DataFrame): DataFrame with SharePoint items
        period (dict, optional): Dictionary with filter parameters:
            - 'type': 'month', 'year', or 'all'
            - 'year': Year to filter for (int)
            - 'month': Month to filter for (int, 1-12)
    """
    if df.empty:
        return {"total_users": 0, "total_tasks": 0, "tasks_completed": 0, "tasks_pending": 0, "tasks_missed": 0, "orders_received": 0, "changes": {}}
    
    uae_tz = pytz.timezone("Asia/Dubai")
    now_uae = datetime.now(uae_tz)
    
    # Convert dates to UAE timezone
    df['BCD'] = pd.to_datetime(df['BCD'], errors='coerce', utc=True).dt.tz_convert(uae_tz)
    df['Created'] = pd.to_datetime(df['Created'], errors='coerce', utc=True).dt.tz_convert(uae_tz)

    # Apply period filter if specified
    if period and period['type'] != 'all':
        filtered_df = df.copy()
        if period['type'] == 'month':
            filtered_df = df[
                (df['Created'].dt.year == period['year']) & 
                (df['Created'].dt.month == period['month'])
            ]
        elif period['type'] == 'year':
            filtered_df = df[df['Created'].dt.year == period['year']]
    else:
        filtered_df = df

    total_users = filtered_df['AssignedTo'].nunique()
    total_tasks = len(filtered_df)
    tasks_completed = len(filtered_df[filtered_df['SubmissionStatus'] == 'Submitted'])
    tasks_pending = len(filtered_df[(filtered_df['SubmissionStatus'] != 'Submitted') & (filtered_df['BCD'] >= now_uae)])
    tasks_missed = len(filtered_df[(filtered_df['SubmissionStatus'] != 'Submitted') & (filtered_df['BCD'] < now_uae)])
    orders_received = len(filtered_df[filtered_df['Status'] == 'Received']) if 'Status' in filtered_df.columns else 0

    # Calculate month-over-month changes if viewing current month
    changes = {}
    if period and period['type'] == 'month':
        # Get last month's data
        last_month = period['month'] - 1
        last_year = period['year']
        if last_month == 0:
            last_month = 12
            last_year -= 1

        last_month_df = df[
            (df['Created'].dt.year == last_year) & 
            (df['Created'].dt.month == last_month)
        ]

        if not last_month_df.empty:
            last_month_stats = {
                "total_tasks": len(last_month_df),
                "tasks_completed": len(last_month_df[last_month_df['SubmissionStatus'] == 'Submitted']),
                "orders_received": len(last_month_df[last_month_df['Status'] == 'Received']) if 'Status' in last_month_df.columns else 0
            }

            changes = {
                "total_tasks_change": ((total_tasks - last_month_stats["total_tasks"]) / last_month_stats["total_tasks"] * 100) if last_month_stats["total_tasks"] > 0 else 0,
                "completed_tasks_change": ((tasks_completed - last_month_stats["tasks_completed"]) / last_month_stats["tasks_completed"] * 100) if last_month_stats["tasks_completed"] > 0 else 0,
                "orders_received_change": ((orders_received - last_month_stats["orders_received"]) / last_month_stats["orders_received"] * 100) if last_month_stats["orders_received"] > 0 else 0
            }

    return {
        "total_users": total_users,
        "total_tasks": total_tasks,
        "tasks_completed": tasks_completed,
        "tasks_pending": tasks_pending,
        "tasks_missed": tasks_missed,
        "orders_received": orders_received,
        "changes": changes
    }
    
        
def compute_user_analytics(df):
    if df.empty or 'AssignedTo' not in df.columns:
        return {}

    uae_tz = pytz.timezone("Asia/Dubai")
    now_uae = datetime.now(uae_tz)

    # Convert dates to timezone-aware
    if 'BCD' in df.columns:
        df['BCD'] = pd.to_datetime(df['BCD'], errors='coerce', utc=True).dt.tz_convert(uae_tz)
    if 'Start Date' in df.columns:
        df['Start Date'] = pd.to_datetime(df['Start Date'], errors='coerce', utc=True).dt.tz_convert(uae_tz)

    analytics = {}
    for user, user_df in df.groupby('AssignedTo'):
        # Last assigned date = latest Start Date
        last_assigned_date = None
        if 'Start Date' in user_df.columns and not user_df['Start Date'].isna().all():
            last_assigned_date = user_df['Start Date'].max()
            last_assigned_date = last_assigned_date.strftime("%Y-%m-%d %H:%M")

        analytics[user] = {
            "total_tasks": len(user_df),
            "tasks_completed": len(user_df[user_df['SubmissionStatus']=='Submitted']),
            "tasks_pending": len(user_df[(user_df['SubmissionStatus']!='Submitted') & (user_df['BCD']>=now_uae)]),
            "tasks_missed": len(user_df[(user_df['SubmissionStatus']!='Submitted') & (user_df['BCD']<now_uae)]),
            "orders_received": len(user_df[user_df['Status']=='Received']) if 'Status' in df.columns else 0,
            "last_assigned_date": last_assigned_date
        }

    return analytics



def compute_user_analytics_with_last_date(df, EXCLUDED_USERS, period=None):
    """
    Compute user analytics with optional period filtering.
    """

    if df.empty or 'AssignedTo' not in df.columns:
        return {}

    # ✅ Make a copy to avoid SettingWithCopyWarning
    df = df[~df['AssignedTo'].isin(EXCLUDED_USERS)].copy()
    if df.empty:
        return {}

    uae_tz = pytz.timezone("Asia/Dubai")
    now_uae = datetime.now(uae_tz)

    # ✅ Safely update datetime fields
    df['Created'] = pd.to_datetime(df['Created'], errors='coerce', utc=True).dt.tz_convert(uae_tz)
    df['BCD'] = pd.to_datetime(df['BCD'], errors='coerce', utc=True).dt.tz_convert(uae_tz)

    # Apply period filter if specified
    if period and period['type'] != 'all':
        if period['type'] == 'month':
            df = df[
                (df['Created'].dt.year == period['year']) &
                (df['Created'].dt.month == period['month'])
            ].copy()
        elif period['type'] == 'year':
            df = df[df['Created'].dt.year == period['year']].copy()

    # Detect Start Date column (case/space tolerant)
    start_col = next((col for col in df.columns if col.lower().replace(" ", "") == "startdate"), None)

    if start_col:
        df[start_col] = pd.to_datetime(df[start_col], errors='coerce', utc=True).dt.tz_convert(uae_tz)

    analytics = {}
    for user, user_df in df.groupby('AssignedTo'):
        last_assigned = user_df[start_col].max() if start_col else None

        analytics[user] = {
            "total_tasks": len(user_df),
            "tasks_completed": len(user_df[user_df['SubmissionStatus'] == 'Submitted']),
            "tasks_pending": len(user_df[(user_df['SubmissionStatus'] != 'Submitted') & (user_df['BCD'] >= now_uae)]),
            "tasks_missed": len(user_df[(user_df['SubmissionStatus'] != 'Submitted') & (user_df['BCD'] < now_uae)]),
            "orders_received": len(user_df[user_df['Status'] == 'Received']) if 'Status' in df.columns else 0,
            "last_assigned_date": last_assigned.strftime("%Y-%m-%d %H:%M") if pd.notna(last_assigned) else None
        }

    return analytics


def extract_usernames_from_df(df, user_columns=None, exclude_users=None):
    """
    Extract all unique usernames from a DataFrame containing SharePoint items,
    with an option to exclude certain users.
    
    Args:
        df (pd.DataFrame): DataFrame with SharePoint list items.
        user_columns (list, optional): List of column names to extract usernames from. 
                                       Defaults to ['AssignedTo', 'Author', 'Editor'].
        exclude_users (list, optional): List of usernames to exclude. Defaults to None.
    
    Returns:
        set: Set of unique usernames excluding the specified users.
    """
    if user_columns is None:
        user_columns = ['AssignedTo', 'Author', 'Editor']
    if exclude_users is None:
        exclude_users = []

    usernames = set()
    for col in user_columns:
        if col in df.columns:
            df[col].dropna().apply(
                lambda x: [usernames.add(u.strip()) for u in str(x).split(',') if u.strip() not in exclude_users]
            )

    return usernames


def get_user_details(access_token, usernames):
    """
    Fetch details of users from Microsoft Graph API given a list of usernames (UPNs).
    
    Returns a list of dicts with user information.
    """
    headers = {"Authorization": f"Bearer {access_token}"}
    user_details = []

    for username in usernames:
        # Microsoft Graph API: Get user by UPN/email
        url = f"https://graph.microsoft.com/v1.0/users/{username}"
        resp = requests.get(url, headers=headers)
        if resp.status_code == 200:
            user_details.append(resp.json())
        else:
            print(f"Failed to fetch {username}: {resp.text}")

    return user_details


def generate_user_analytics(df, user_column='AssignedTo', status_column='Status', 
                            title_column='Title', due_column='BCD', start_column='StartDate',
                            assigned_column='AssignedDate', orders_column='OrdersReceived',
                            exclude_users=None):
    """
    Generate per-user analytics with counts, lists of tasks, last assigned date,
    and orders received, excluding specific users. Considers tasks with future start dates as ongoing.
    """
    if exclude_users is None:
        exclude_users = []

    analytics = []

    if df.empty:
        return pd.DataFrame(analytics)

    # Ensure date columns are datetime with UTC timezone
    df[due_column] = pd.to_datetime(df[due_column], utc=True, errors='coerce')
    df[start_column] = pd.to_datetime(df[start_column], utc=True, errors='coerce')
    
    if assigned_column in df.columns:
        df[assigned_column] = pd.to_datetime(df[assigned_column], utc=True, errors='coerce')
    else:
        df[assigned_column] = df[start_column]  # fallback if no assigned date column

    now_utc = pd.Timestamp.now(tz='UTC')

    # Filter out excluded users
    df_filtered = df[~df[user_column].isin(exclude_users)]

    # Group by user
    grouped = df_filtered.groupby(user_column)

    for user, group in grouped:
        total_tasks = len(group)

        # Completed tasks
        completed_tasks_df = group[group[status_column] == 'Completed']

        # Ongoing tasks: Not submitted + either due date in future or start date in future
        ongoing_tasks_df = group[
            (group[status_column] != 'Submitted') & 
            ((group[due_column] >= now_utc) | (group[start_column] >= now_utc))
        ]

        # Missed tasks: Not submitted + due date in past
        missed_tasks_df = group[
            (group[status_column] != 'Submitted') & 
            (group[due_column] < now_utc)
        ]

        # ✅ Last assigned date: use AssignedDate if exists
        last_assigned_date = group[assigned_column].max() if not group[assigned_column].dropna().empty else None

        # Orders received
        orders_received = group[orders_column].sum() if orders_column in group.columns else 0

        analytics.append({
            'User': user,
            'TotalTasks': total_tasks,
            'CompletedTasksCount': len(completed_tasks_df),
            'OngoingTasksCount': len(ongoing_tasks_df),
            'MissedTasksCount': len(missed_tasks_df),
            'CompletedTasks': completed_tasks_df[title_column].tolist(),
            'OngoingTasks': ongoing_tasks_df[title_column].tolist(),
            'MissedTasks': missed_tasks_df[title_column].tolist(),
            'LastAssignedDate': last_assigned_date.isoformat() if last_assigned_date is not None else None,
            'OrdersReceived': orders_received
        })

    analytics_df = pd.DataFrame(analytics)
    return analytics_df


def get_user_analytics_specific(df: pd.DataFrame, username: str) -> dict:
    """
    Returns analytics and task lists for a specific user.
    """
    if df.empty:
        return {
            'Username': username,
            'TotalTasks': 0,
            'OngoingTasksCount': 0,
            'CompletedTasksCount': 0,
            'MissedTasksCount': 0,
            'OngoingTasks': [],
            'CompletedTasks': [],
            'MissedTasks': []
        }

    # Ensure DueDate is datetime with UTC timezone
    df['DueDate'] = pd.to_datetime(df['DueDate'], utc=True)

    # Safe UTC now
    now_utc = pd.Timestamp.utcnow()
    if now_utc.tzinfo is None:
        now_utc = now_utc.tz_localize('UTC')
    else:
        now_utc = now_utc.tz_convert('UTC')

    # Filter tasks for the user
    user_tasks = df[df['AssignedTo'] == username]

    completed_tasks = user_tasks[user_tasks['Status'] == 'Completed']

    ongoing_tasks = user_tasks[
        (user_tasks['Status'] != 'Submitted') &
        (user_tasks['DueDate'] >= now_utc)
    ]

    missed_tasks = user_tasks[
        (user_tasks['Status'] != 'Submitted') &
        (user_tasks['DueDate'] < now_utc)
    ]

    return {
        'Username': username,
        'TotalTasks': len(user_tasks),
        'OngoingTasksCount': len(ongoing_tasks),
        'CompletedTasksCount': len(completed_tasks),
        'MissedTasksCount': len(missed_tasks),
        'OngoingTasks': ongoing_tasks.to_dict('records'),
        'CompletedTasks': completed_tasks.to_dict('records'),
        'MissedTasks': missed_tasks.to_dict('records'),
        'OrdersReceived': len(user_tasks[user_tasks['Order Status']=='Received']) if 'Order status' in user_tasks.columns else 0 ,
        'user_tasks': user_tasks.to_dict('records')
    }


# ==============================================================================
# ==||| FOR BUSINESS CARDS |||==
# ==============================================================================


def get_onedrive_access_token():
    """Acquires an access token for OneDrive operations."""
    result = onedrive_msal_app.acquire_token_silent(SCOPE_ONEDRIVE, account=None)
    if not result:
        result = onedrive_msal_app.acquire_token_for_client(scopes=SCOPE_ONEDRIVE)
    if "access_token" in result:
        return result['access_token']
    else:
        raise Exception(f"Failed to acquire OneDrive token: {result.get('error_description')}")

def get_all_contacts_from_onedrive():
    """Fetches all data from the Contacts.xlsx file in the specified user's OneDrive."""
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}"}
        
        url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_USER_ID}/drive/root:/"
            f"{FILE_PATH}:/workbook/worksheets('{CONTACT_WORKSHEET_NAME}')/usedRange"
        )
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        rows = data.get('values', [])
        if not rows or len(rows) < 2:
            return [] 

        header = rows[0]
        contacts = []
        for i, row_data in enumerate(rows[1:]):
            contact_dict = {header[j]: row_data[j] if j < len(row_data) else "" for j in range(len(header))}
            contact_dict['row_id'] = i + 2  # Excel rows are 1-based, data starts on row 2
            contacts.append(contact_dict)
        return contacts
    except Exception as e:
        print(f"Error fetching contacts from OneDrive: {e}")
        return []

def update_contact_in_onedrive_excel(row_id, updated_data_dict):
    """Updates a single row in the Contacts.xlsx file."""
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}"}
        
        header_url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_USER_ID}/drive/root:/"
            f"{FILE_PATH}:/workbook/worksheets('{CONTACT_WORKSHEET_NAME}')/range(address='A1:Z1')"
        )
        header_res = requests.get(header_url, headers=headers)
        header_res.raise_for_status()
        header = header_res.json().get("values", [[]])[0]
        if not header: raise Exception("Could not retrieve header row.")
        
        values_to_update = [updated_data_dict.get(col_name, "") for col_name in header]
        last_col = chr(ord('A') + len(header) - 1)
        range_address = f"A{row_id}:{last_col}{row_id}"
        
        update_url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_USER_ID}/drive/root:/"
            f"{FILE_PATH}:/workbook/worksheets('{CONTACT_WORKSHEET_NAME}')/range(address='{range_address}')"
        )
        
        patch_res = requests.patch(update_url, headers=headers, json={"values": [values_to_update]})
        patch_res.raise_for_status()
        return True
    except Exception as e:
        print(f"Error updating contact in OneDrive: {e}")
        return False


# ==============================================================================
# ==============================================================================

def get_all_customers_from_onedrive():
    """Fetches all data from the Customers.xlsx file in the specified user's OneDrive."""
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}"}
        
        url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/"
            f"Customers.xlsx:/workbook/worksheets('Sheet1')/usedRange"
        )
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        rows = data.get('values', [])
        if not rows or len(rows) < 2:
            return [] 

        header = rows[0]
        customers = []
        for i, row_data in enumerate(rows[1:]):
            customer_dict = {header[j]: row_data[j] if j < len(row_data) else "" for j in range(len(header))}
            customer_dict['row_id'] = i + 2  # Excel rows are 1-based, data starts on row 2
            customers.append(customer_dict)
        return customers
    except Exception as e:
        print(f"Error fetching customers from OneDrive: {e}")
        return []
    
    
def get_user_details_from_excell():
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}"}
        
        url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/"
            f"Userdatas.xlsx:/workbook/worksheets('Sheet1')/usedRange"
        )
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        rows = data.get('values', [])
        if not rows or len(rows) < 2:
            return [] 

        header = rows[0]
        customers = []
        for i, row_data in enumerate(rows[1:]):
            customer_dict = {header[j]: row_data[j] if j < len(row_data) else "" for j in range(len(header))}
            customer_dict['row_id'] = i + 2  # Excel rows are 1-based, data starts on row 2
            customers.append(customer_dict)
        return customers
    except Exception as e:
        print(f"Error fetching customers from OneDrive: {e}")
        return []
    

    
def get_user_tasks_details_from_excell():
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}"}
        
        url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/"
            f"Sharepoint Datas.xlsx:/workbook/worksheets('Sheet1')/usedRange"
        )
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        
        rows = data.get('values', [])
        if not rows or len(rows) < 2:
            return [] 

        header = rows[0]
        customers = []
        for i, row_data in enumerate(rows[1:]):
            customer_dict = {header[j]: row_data[j] if j < len(row_data) else "" for j in range(len(header))}
            customer_dict['row_id'] = i + 2  # Excel rows are 1-based, data starts on row 2
            customers.append(customer_dict)
        return customers
    except Exception as e:
        print(f"Error fetching customers from OneDrive: {e}")
        return []
    
    
def upload_photo_to_onedrive(photo_file, user_id, email):
    """
    Uploads a user profile photo to OneDrive and returns the shared link.
    Uses Graph API user ID for unique naming.
    
    photo_file: Werkzeug FileStorage (from Flask request.files)
    user_id: Graph API user id
    email: used for readability in filename
    """
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}"}

        # Extract file extension
        file_ext = photo_file.filename.split('.')[-1]
        filename = f"profile_photos/{email}_{user_id}.{file_ext}"

        # Read file content
        file_content = photo_file.read()

        # Upload file to OneDrive
        upload_url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/{filename}:/content"
        )
        response = requests.put(upload_url, headers=headers, data=file_content)
        response.raise_for_status()
        uploaded_file = response.json()

        # Create a shareable link
        link_url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/items/{uploaded_file['id']}/createLink"
        )
        payload = {"type": "view", "scope": "anonymous"}  # anonymous view link
        link_response = requests.post(link_url, headers=headers, json=payload)
        link_response.raise_for_status()

        share_link = link_response.json()['link']['webUrl']
        return share_link

    except Exception as e:
        print(f"Error uploading photo to OneDrive: {e}")
        return ""

    
def add_or_update_user_in_excel(email, user_id, name, role, photo_file=None):
    """
    Adds a new user or updates an existing user in Excel.
    Uses Graph API user ID as unique identifier.
    """
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

        users = get_user_details_from_excell()
        user = next((u for u in users if u.get("email", "").lower() == email.lower()), None)

        # Upload photo to OneDrive if provided
        dp_url = ""
        if photo_file:
            dp_url = upload_photo_to_onedrive(photo_file, user_id, email)

        if user:
            # Update existing row
            row_id = user["row_id"]
            update_values = [[name, email, role, dp_url, 1]]  # columns: name, email, role, dp_url, flag
            update_url = f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/Userdatas.xlsx:/workbook/worksheets('Sheet1')/range(address='A{row_id}:E{row_id}')"
            response = requests.patch(update_url, headers=headers, json={"values": update_values})
            response.raise_for_status()
        else:
            # Append new row (Graph API user ID as unique identifier)
            append_values = [[user_id, name, email, role, dp_url, 1]]  # columns: user_id, name, email, role, dp_url, flag
            append_url = f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/Userdatas.xlsx:/workbook/worksheets('Sheet1')/tables('Table1')/rows/add"
            response = requests.post(append_url, headers=headers, json={"values": append_values})
            response.raise_for_status()

        return True

    except Exception as e:
        print(f"Error adding/updating user in Excel: {e}")
        return False

        
# ==============================================================================
# ==============================================================================


def get_task_details(df: pd.DataFrame, task_title: str) -> dict:
    """
    Returns details of a specific task by title.
    """
    if df.empty:
        return {}

    task_row = df[df['Title'] == task_title]
    if task_row.empty:
        return {}

    return task_row.iloc[0].to_dict()


DOMAIN=os.getenv("DOMAIN")
# --- Send quote approval email ---


def send_quote_approval_email(quote_data, submitter_email, admin_emails):
    token = get_access_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    # ✅ 1️⃣ Create Adaptive Card (goes HERE)
    adaptive_card = {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": [
            {"type": "TextBlock", "text": "New Quote Submission", "weight": "Bolder", "size": "Medium"},
            {"type": "TextBlock", "text": f"Quote Reference: {quote_data['reference'][0]}", "wrap": True},
        ],
        "actions": [
            {
                "type": "Action.Http",
                "title": "Approve",
                "method": "POST",
                "url": f"https://{DOMAIN}/quote_decision",
                "body": json.dumps({
                    "decision": "approve",
                    "quote_reference": quote_data['reference'][0],
                    "submitter_email": submitter_email
                }),
                "headers": [{"name": "Content-Type", "value": "application/json"}],
                "authentication": {"type": "None"}
            },
            {
                "type": "Action.Http",
                "title": "Reject",
                "method": "POST",
                "url": f"https://{DOMAIN}/quote_decision",
                "body": json.dumps({
                    "decision": "reject",
                    "quote_reference": quote_data['reference'][0],
                    "submitter_email": submitter_email
                }),
                "headers": [{"name": "Content-Type", "value": "application/json"}],
                "authentication": {"type": "None"}
            }
        ]
    }

    # ✅ 2️⃣ Email payload (comes after adaptive_card)
    message = {
        "message": {
            "subject": f"Quote Approval Required - {quote_data['reference'][0]}",
            "body": {
                "contentType": "html",
                "content": f"""
                <html>
                <body>
                    <p>New quote submitted by {submitter_email}</p>
                    <p>Quote Reference: {quote_data['reference'][0]}</p>
                    <script type="application/adaptivecard+json">
                        {json.dumps(adaptive_card)}
                    </script>
                </body>
                </html>
                """
            },
            "toRecipients": [{"emailAddress": {"address": admin}} for admin in admin_emails]
        }
    }

    # ✅ 3️⃣ Send the mail
    url = f"{GRAPH_API_ENDPOINT}/users/{submitter_email}/sendMail"
    response = requests.post(url, headers=headers, json=message)
    response.raise_for_status()

    print("Approval email sent successfully")
    return True


def add_sharepoint_list_item(item_fields):
    token = get_access_token()
    site_id = get_site_id(token, "hamdaz1.sharepoint.com", "/sites/Test")
    list_id = get_list_id(token, site_id, "Quotes")

    url = f"{GRAPH_API_ENDPOINT}/sites/{site_id}/lists/{list_id}/items"
    payload = {"fields": item_fields}
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    try:
        resp = requests.post(url, headers=headers, json=payload)
        resp.raise_for_status()
        print(f"✅ Item {item_fields.get('Reference')} added successfully")
        return resp.json()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error adding item {item_fields.get('Reference')} to SharePoint: {e}")
        return None



from urllib.parse import quote





def update_sharepoint_item(reference, update_fields):
    """Update SharePoint item based on unique reference field."""
    token = get_access_token()
    site_id = get_site_id(token, "hamdaz1.sharepoint.com", "/sites/Test")
    list_id = get_list_id(token, site_id, "Quotes")

    # Search for item with the reference
    url = f"{GRAPH_API_ENDPOINT}/sites/{site_id}/lists/{list_id}/items?filter=fields/Reference eq '{reference}'"
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    items = resp.json().get("value", [])
    if not items:
        raise ValueError(f"No SharePoint item found with Reference {reference}")

    item_id = items[0]["id"]
    update_url = f"{GRAPH_API_ENDPOINT}/sites/{site_id}/lists/{list_id}/items/{item_id}/fields"
    resp = requests.patch(update_url, headers=headers, json=update_fields)
    resp.raise_for_status()
    return resp.json()



import requests

def get_list_columns(site_domain, site_path, list_name):
    token = get_access_token()
    site_id = get_site_id(token, site_domain, site_path)
    list_id = get_list_id(token, site_id, list_name)
    
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/columns"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    columns = response.json().get("value", [])
    # for col in columns:
    #     print(f"Display Name: {col['displayName']}, Internal Name: {col['name']}")
    return columns




def generate_sharepoint_filter_endpoint(site_domain, site_path, list_name, columns, user_prompt):
    """
    Ask the LLM to generate a SharePoint API endpoint with $filter query
    based on the columns and the user's natural language prompt.

    Returns:
        str: Graph API endpoint URL (relative to site) with filters.
    """

    system_msg = (
        "You are an expert at generating Microsoft Graph API queries for SharePoint lists. "
        "Columns available: " + ", ".join(columns) + ".\n"
        "Given the user's request, generate the endpoint URL for fetching the list items with proper $filter parameters. "
        "Return only the URL string, do not include extra text.\n"
        "Use ISO 8601 dates where needed, and use logical operators eq, ge, le, etc. "
        "Example: /sites/{site_id}/lists/{list_id}/items?$filter=Status eq 'Pending' and AssignedTo eq 'John'"
    )

    messages = [
        {"role": "system", "content": system_msg},
        {"role": "user", "content": user_prompt}
    ]

    response = openai.ChatCompletion.create(
        model="gpt-4o",
        messages=messages,
        temperature=0,
        max_tokens=2000
    )

    endpoint_url = response.choices[0].message.content.strip()
    return endpoint_url




def fetch_filtered_sharepoint_data(access_token, endpoint_url):
    """
    Calls the LLM-generated endpoint to fetch filtered SharePoint list items.
    """
    base_url = "https://graph.microsoft.com/v1.0"
    full_url = f"{base_url}{endpoint_url}"

    headers = {"Authorization": f"Bearer {access_token}"}
    resp = requests.get(full_url, headers=headers)
    resp.raise_for_status()
    items = resp.json().get("value", [])
    return items






def ensure_sharepoint_folder(access_token, site_id, library_name, folder_path):
    """
    Ensure a folder exists in a SharePoint document library.
    folder_path example: 'QuoteCostingSheets/QT-1001'
    Returns the folder ID.
    """
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"{GRAPH_API_ENDPOINT}/sites/{site_id}/drive/root:/{folder_path}"
    resp = requests.get(url, headers=headers)
    
    if resp.status_code == 404:
        # Folder does not exist, create it
        parent_path, folder_name = folder_path.rsplit('/', 1)
        create_url = f"{GRAPH_API_ENDPOINT}/sites/{site_id}/drive/root:/{parent_path}:/children"
        payload = {
            "name": folder_name,
            "folder": {},
            "@microsoft.graph.conflictBehavior": "rename"
        }
        resp = requests.post(create_url, headers={**headers, "Content-Type": "application/json"}, json=payload)
        resp.raise_for_status()
        return resp.json()["id"]
    else:
        resp.raise_for_status()
        return resp.json()["id"]

def upload_file_to_sharepoint(access_token, site_id, folder_path, file_name, file_bytes):
    """
    Upload file to SharePoint folder.
    """
    url = f"{GRAPH_API_ENDPOINT}/sites/{site_id}/drive/root:/{folder_path}/{file_name}:/content"
    headers = {"Authorization": f"Bearer {access_token}"}
    resp = requests.put(url, headers=headers, data=file_bytes)
    resp.raise_for_status()
    return resp.json()["webUrl"]  # Download link



def generate_quote_excel(quote):
    """
    Generates an Excel file for a single quote with all items and metadata.
    Saves to QuoteCostingSheets folder.
    """
    folder_path = os.path.join("QuoteCostingSheets")
    os.makedirs(folder_path, exist_ok=True)
    reference = quote.get("Reference", f"Quote_{datetime.now().strftime('%Y%m%d%H%M%S')}")
    file_path = os.path.join(folder_path, f"{reference}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote Items"

    # Columns to include (rows = items)
    columns = [
        "Quote Date","Expiry Date","Quote Number","Quote Status","Customer Name",
        "VAT Treatment","Place Of Supply","Is Inclusive Tax","Project Name","Project ID",
        "PurchaseOrder","Currency Code","Exchange Rate","Discount Type","Is Discount Before Tax",
        "Entity Discount Percent","Entity Discount Amount","Item Name","SKU","Account",
        "Item Desc","Tax Registration Number","Quantity","Usage unit","Item Price",
        "Discount","Discount Amount","Item Tax","Item Tax %","Item Tax Type",
        "Out of Scope Reason","Item Tax Exemption Reason","Item Type","Template Name",
        "Sales person","Notes","Terms & Conditions"
    ]

    ws.append(columns)

    # Load items
    items = json.loads(quote.get("AllItems","[]"))

    for item in items:
        row = [
            quote.get("QuoteDate"),
            quote.get("ExpiryDate"),
            quote.get("Reference"),
            quote.get("ApprovalStatus"),
            quote.get("CustomerID"),
            quote.get("TaxTreatment"),
            "", "", "", "",
            "", quote.get("Currency"), "", "", "",
            "", "",
            item.get("ItemDetails"),
            item.get("Brand"),
            "",
            "", item.get("Quantity"), "",
            item.get("Rate"),
            "", "", item.get("Tax"),
            "", "", "",
            "", item.get("ItemType",""),
            "", quote.get("QuoteCreator"),
            quote.get("CustomerNotes",""),
            quote.get("TermsConditions","")
        ]
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)
    return file_path


import requests

def get_excel_data_from_onedrive(file_name, sheet_name):
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}"}

        url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/"
            f"{file_name}:/workbook/worksheets('{sheet_name}')/usedRange"
        )

        print(f"Fetching from: {url}")  # 👈 check exact path
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()

        print("Raw data from Graph API:\n", data)  # 👈 see what Excel actually returned

        rows = data.get("values", [])
        if not rows:
            print("⚠️ No rows returned. Possibly wrong file or sheet name.")
            return []

        header = rows[0]
        # customers = []
        # for i, row_data in enumerate(rows[1:]):
        #     row_dict = {header[j]: row_data[j] if j < len(row_data) else "" for j in range(len(header))}
        #     row_dict["row_id"] = i + 2
        #     customers.append(row_dict)

        return data

    except Exception as e:
        print(f"❌ Error fetching Excel data: {e}")
        return []




def get_user_profile_photo():
    access_token = get_access_token()
    url ="https://graph.microsoft.com/v1.0/me/photo/$value"
    headers = {"Authorization": f"Bearer {access_token}"}
    resp = requests.get(url, headers=headers)
    return resp




    



def upload_file_to_sharepoint_folder(folder_path, file_name, file_bytes):
    token = get_access_token()
    site_id = get_site_id(token, "hamdaz1.sharepoint.com", "/sites/Test")
 
    upload_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{folder_path}/{file_name}:/content"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    resp = requests.put(upload_url, headers=headers, data=file_bytes)
    resp.raise_for_status()
    share_link = resp.json()["webUrl"]
    return share_link # returns metadata including 'id', 'webUrl' etc.



def update_sharepoint_item_with_link(item_id, link_url):
    access_token = get_access_token()
    site_id = get_site_id(access_token, "hamdaz1.sharepoint.com", "/sites/Test")
    list_id = get_list_id(access_token, site_id, "Quotes")

    patch_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items/{item_id}/fields"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    # ✅ The correct way to update a Hyperlink column via Graph API
    data = {
        "attachmentlink": f"{link_url}"
    }

    resp = requests.patch(patch_url, headers=headers, json=data)
    resp.raise_for_status()
    return resp.json()


# ============================================================================


def update_user_analytics_in_sharepoint(item_id, item_fields):
    access_token = get_access_token()
    site_id = get_site_id(access_token, "hamdaz1.sharepoint.com", "/sites/Test")
    list_id = get_list_id(access_token, site_id, "useranalytics")

    patch_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/items/{item_id}/fields"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    # ✅ Send fields dictionary directly
    resp = requests.patch(patch_url, headers=headers, json=item_fields)
    resp.raise_for_status()
    return resp.json()




def add_item_to_sharepoint(item_fields):
    token = get_access_token()
    site_id = get_site_id(token, "hamdaz1.sharepoint.com", "/sites/Test")
    list_id = get_list_id(token, site_id, "useranalytics")

    url = f"{GRAPH_API_ENDPOINT}/sites/{site_id}/lists/{list_id}/items"
    payload = {"fields": item_fields}
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    try:
        resp = requests.post(url, headers=headers, json=payload)
        resp.raise_for_status()
        print(f"✅ Item {item_fields.get('Username')} added successfully")
        return resp.json()
    except requests.exceptions.RequestException as e:
        print(f"❌ Error adding item {item_fields.get('Username')} to SharePoint: {e}")
        return None


def get_existing_useranalytics_items():
    token = get_access_token()
    site_id = get_site_id(token, "hamdaz1.sharepoint.com", "/sites/Test")
    list_id = get_list_id(token, site_id, "useranalytics")
    url = f"{GRAPH_API_ENDPOINT}/sites/{site_id}/lists/{list_id}/items?expand=fields"
    headers = {"Authorization": f"Bearer {token}"}

    all_items = []

    while url:
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        data = resp.json()
        all_items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")  # Fetch next page if exists

    return all_items



from datetime import datetime, timezone

from datetime import datetime, timezone

from datetime import datetime, timezone

def calculate_priority_score(user_analytics):
    """
    Calculate priority score based on:
    1. Low active tasks → higher priority
    2. Among similar tasks, longer idle → higher priority
    """
    now = datetime.now(timezone.utc)
    scores = []

    for _, row in user_analytics.iterrows():
        active_tasks = int(row["OngoingTasksCount"])
        last_assigned = row["LastAssignedDate"]

        if isinstance(last_assigned, str):
            last_assigned = datetime.fromisoformat(last_assigned)

        days_since_last = (now - last_assigned).total_seconds() / (24 * 3600)

        # Primary: negative active tasks (so fewer tasks → higher score)
        # Secondary: days_since_last (more idle → higher score)
        score = (-active_tasks, days_since_last)
        scores.append(score)

    df = user_analytics.copy()
    df["PriorityScore"] = scores
    return df


def assign_priority_rank(user_analytics):
    """
    Assign priority rank based on tuple sorting:
    - Highest priority: lowest active tasks, then longest idle
    """
    df = user_analytics.copy()

    # Sort by tuple: (-active_tasks, days_since_last)
    df = df.sort_values(by="PriorityScore", ascending=False).reset_index(drop=True)

    # Assign rank
    df["PriorityRank"] = df.index + 1
    return df



def find_existing_user_item(existing_items, username):
    """Find existing SharePoint row matching the username (case-insensitive)."""
    username = username.lower().strip()

    for item in existing_items:
        fields = item.get("fields", {})
        normalized = {k.lower(): str(v).lower() for k, v in fields.items()}
        if normalized.get("username") == username or username in normalized.values():
            return item

    return None

def list_org_users(access_token):
    """
    List all users in the organization using Microsoft Graph API.

    Args:
        access_token (str): Azure AD access token with User.Read.All permission.

    Returns:
        list: List of users with basic info (id, displayName, mail, userPrincipalName)
    """
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    users = []
    url = f"{GRAPH_API_ENDPOINT}/users"
    
    while url:
        response = requests.get(url, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Error fetching users: {response.status_code} - {response.text}")
        
        data = response.json()
        users.extend(data.get("value", []))
        
        # Pagination: check if there is a next page
        url = data.get("@odata.nextLink", None)

    # Return simplified list
    return [
        {
            "id": u.get("id"),
            "displayName": u.get("displayName"),
            "mail": u.get("mail"),
            "userPrincipalName": u.get("userPrincipalName")
        }
        for u in users
    ]




def get_partnership_data():
    """Fetches data from competitor_contact_info_mock.xlsx in OneDrive via Graph API."""
    try:
        access_token = get_onedrive_access_token()
        headers = {"Authorization": f"Bearer {access_token}"}

        url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/"
            f"competitor_contact_info_full.xlsx:/workbook/worksheets('Sheet1')/usedRange"
        )

        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()

        rows = data.get('values')
        if not rows or len(rows) < 2:
            print("[WARN] No valid rows found in Excel file.")
            return []

        headers_row = rows[0]
        records = rows[1:]

        # Safeguard in case headers are missing or invalid
        if not isinstance(headers_row, list):
            print("[ERROR] Invalid header format in Excel.")
            return []

        df = pd.DataFrame(records, columns=headers_row)
        df = df.fillna("")
        return df.to_dict(orient="records")

    except Exception as e:
        print(f"[ERROR] Failed to fetch or parse OneDrive Excel: {e}")
        return []  # Return an empty list instead of None
    


def save_partnership_update(product_group, product_name, manufacturer, competitor_name, field, new_value):
    """
    Updates only one cell (Status or Remarks) in the Excel file on OneDrive.
    Matches based on Product Group Number, Product, and Manufacturer.
    """
    try:
        access_token = get_onedrive_access_token()
        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        # 1️⃣ Fetch the Excel data
        read_url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/"
            f"competitor_contact_info_full.xlsx:"
            f"/workbook/worksheets('Sheet1')/usedRange"
        )
        read_resp = requests.get(read_url, headers=headers)
        read_resp.raise_for_status()
        read_values = read_resp.json().get('values', [])

        if not read_values or len(read_values) < 2:
            print("[ERROR] No data found in Excel file.")
            return False

        headers_row = read_values[0]
        data_rows = read_values[1:]

        # 2️⃣ Get the column index
        if field not in headers_row:
            print(f"[ERROR] Column '{field}' not found in Excel.")
            return False

        col_idx = headers_row.index(field)

        # 3️⃣ Find matching row
        row_idx = None
        for i, row in enumerate(data_rows, start=2):  # +1 for header
            try:
                if (
                    str(row[headers_row.index("Product Group Number")]).strip() == str(product_group).strip()
                    and str(row[headers_row.index("Product")]).strip() == str(product_name).strip()
                    and str(row[headers_row.index("ADNOC Approved Manufacturer")]).strip() == str(manufacturer).strip()
                    and str(row[headers_row.index("Competitor Company")]).strip() == str(competitor_name).strip()
                ):
                    row_idx = i
                    break
            except Exception:
                continue

        if not row_idx:
            print("[WARN] No matching row found for update.")
            return False

        # 4️⃣ Convert to Excel column letter (A, B, C…)
        def excel_column_letter(n):
            result = ""
            while n > 0:
                n, remainder = divmod(n - 1, 26)
                result = chr(65 + remainder) + result
            return result

        col_letter = excel_column_letter(col_idx + 1)
        cell_address = f"Sheet1!{col_letter}{row_idx}"

        # 5️⃣ Send PATCH update request
        update_url = (
            f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/root:/"
            f"competitor_contact_info_full.xlsx:"
            f"/workbook/worksheets('Sheet1')/range(address='{cell_address}')"
        )

        payload = {"values": [[new_value]]}
        patch_resp = requests.patch(update_url, headers=headers, json=payload)
        patch_resp.raise_for_status()

        print(f"[INFO] Successfully updated {field} for {manufacturer} ({product_name}) → {new_value}")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to update Excel cell: {e}")
        return False

import urllib.parse

# def get_personal_onedrive_folder():
#     """
#     List all files in 'Documents/defualt_docs' in your personal OneDrive.
#     """
#     try:
#         access_token = get_access_token()  # your existing function
#         headers = {"Authorization": f"Bearer {access_token}"}

#         folder_path = "Documents/defualt_docs"
#         folder_path_encoded = urllib.parse.quote(folder_path)

#         url = f"{GRAPH_API_ENDPOINT}/me/drive/root:/{folder_path_encoded}:/children"
#         print(f"Fetching files from: {url}")

#         response = requests.get(url, headers=headers)
#         response.raise_for_status()

#         files = response.json().get("value", [])
#         print(f"Found {len(files)} files in '{folder_path}':")
#         for f in files:
#             print(f" - {f['name']} ({f.get('webUrl')})")

#         return files

#     except Exception as e:
#         print(f"❌ Error fetching folder data: {e}")
#         return []
    


def get_child_files():
    access_token =get_access_token()
    folder_id =  os.getenv("default_folder_id")
    
    if not access_token:
        raise Exception("No access token in session")

    url = f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/items/{folder_id}/children"

    headers = {
        "Authorization": f"Bearer {access_token}"
    }

    res = requests.get(url, headers=headers)
    res.raise_for_status()

    return res.json().get("value", [])




from docx2pdf import convert
import tempfile


def download_docx(file_id):
    """
    Downloads a DOCX file from OneDrive using file ID.
    Returns the path to the temporary DOCX.
    """
    access_token= get_onedrive_access_token()
    
    url = f"https://graph.microsoft.com/v1.0/users/{ONEDRIVE_PRIMARY_USER_ID}/drive/items/{file_id}/content"

    headers = {"Authorization": f"Bearer {access_token}"}

    response = requests.get(url, headers=headers)
    response.raise_for_status()

    # Save to a temporary DOCX
    temp_docx = tempfile.mktemp(suffix=".docx")
    with open(temp_docx, "wb") as f:
        f.write(response.content)

    return temp_docx

# def convert_docx_to_pdf(input_docx):
#     temp_pdf = tempfile.mktemp(suffix=".pdf")
#     convert(input_docx, temp_pdf)
#     return temp_pdf

# import pythoncom
# import tempfile
# from docx2pdf import convert

# def convert_docx_to_pdf(input_docx):
#     # 1. Initialize Windows COM for this Flask thread
#     pythoncom.CoInitialize()
    
#     temp_pdf = tempfile.mktemp(suffix=".pdf")
    
#     try:
#         convert(input_docx, temp_pdf)
#     except Exception as e:
#         raise e
#     finally:
#         # 2. Release resources
#         pythoncom.CoUninitialize()
        
#     return temp_pdf


import io
def generate_quote_excel(quote):
    # 1. Define the columns exactly as they appear in your sample file
    columns = [
        'Quote Date', 'Expiry Date', 'Quote Number', 'Quote Status', 'Customer Name', 
        'VAT Treatment', 'Place Of Supply', 'Is Inclusive Tax', 'Project Name', 
        'Project ID', 'PurchaseOrder', 'Currency Code', 'Exchange Rate', 'Discount Type', 
        'Is Discount Before Tax', 'Entity Discount Percent', 'Entity Discount Amount', 
        'Item Name', 'SKU', 'Account', 'Item Desc', 'Tax Registration Number', 
        'Quantity', 'Usage unit', 'Item Price', 'Discount', 'Discount Amount', 
        'Item Tax', 'Item Tax %', 'Item Tax Type', 'Out of Scope Reason', 
        'Item Tax Exemption Reason', 'Item Type', 'Template Name', 'Sales person', 
        'Notes', 'Terms & Conditions'
    ]
    
    rows = []
    items = quote.get('AllItems_parsed', [])

    # 2. Iterate through items to build rows
    for item in items:
        # Handle potential None values safely
        qty = float(item.get('Quantity', 1) or 1)
        rate = float(item.get('Rate', 0) or 0)
        discount = float(item.get('Discount', 0) or 0)
        tax_percent = float(item.get('Tax', 0) or 0)
        
        row = {
            'Quote Date': str(quote.get('QuoteDate', ''))[:10],
            'Expiry Date': str(quote.get('ExpiryDate', ''))[:10],
            'Quote Number': quote.get('id', ''),  # Or quote.get('QuoteID')
            'Quote Status': quote.get('ApprovalStatus', 'Draft'),
            'Customer Name': quote.get('CustomerName', ''),
            'VAT Treatment': 'vat registered',    # Default value based on sample
            'Place Of Supply': '',
            'Is Inclusive Tax': 'false',          # Default to false
            'Project Name': '',
            'Project ID': '',
            'PurchaseOrder': '',
            'Currency Code': quote.get('Currency', 'USD'),
            'Exchange Rate': 1.0,
            'Discount Type': 'item_level',
            'Is Discount Before Tax': 'true',
            'Entity Discount Percent': 0.0,
            'Entity Discount Amount': 0.0,
            
            # Item Mapping
            'Item Name': item.get('ItemDetails', 'Service'), # Using ItemDetails as Name
            'SKU': item.get('Brand', ''),                    # Mapping Brand to SKU
            'Account': 'Sales',                              # Default Sales Account
            'Item Desc': item.get('ItemDetails', ''),        # Full description
            'Tax Registration Number': '',
            'Quantity': qty,
            'Usage unit': '',
            'Item Price': rate,
            'Discount': discount,
            'Discount Amount': discount,                     # Assuming fixed amount discount
            'Item Tax': 'Standard Rate' if tax_percent > 0 else '',
            'Item Tax %': tax_percent,
            'Item Tax Type': 'ItemAmount' if tax_percent > 0 else '',
            'Out of Scope Reason': '',
            'Item Tax Exemption Reason': '',
            'Item Type': 'service',
            'Template Name': 'Standard Template',
            'Sales person': quote.get('QuoteCreator', ''),
            'Notes': quote.get('Notes', ''),
            'Terms & Conditions': quote.get('PaymentTerms', '')
        }
        rows.append(row)

    # 3. Create DataFrame
    df = pd.DataFrame(rows, columns=columns)
    
    # 4. Write to Excel BytesIO buffer
    output = io.BytesIO()
    # Use 'xlsxwriter' for better formatting options, or default 'openpyxl'
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Estimates')
    
    output.seek(0)
    return output



def fetch_user_planner_tasks():
    access_token = get_access_token()
    headers = {"Authorization": f"Bearer {access_token}"}
    url = "https://graph.microsoft.com/v1.0/me/planner/tasks"
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    tasks = resp.json().get("value", [])
    
    return tasks

def get_user_teams_chats():
    access_token=get_access_token()
    headers = {
        "Authorization": f"Bearer {access_token}"
    }

    # Step 1: Get all chats for the user
    chat_list_url = f"{GRAPH_API_ENDPOINT}/users/{ONEDRIVE_USER_ID}/chats"
    chat_list_response = requests.get(chat_list_url, headers=headers)
    chat_list_response.raise_for_status()
    chats = chat_list_response.json().get("value", [])

    all_chat_messages = {}

    # Step 2: Iterate each chat and fetch messages
    for chat in chats:
        chat_id = chat["id"]
        messages_url = f"{GRAPH_API_ENDPOINT}/chats/{chat_id}/messages"
        messages_response = requests.get(messages_url, headers=headers)
        messages_response.raise_for_status()

        all_chat_messages[chat_id] = messages_response.json().get("value", [])

    return all_chat_messages

# def compare_rfq_and_quote(rfq_text, quote_text):
#     """
#     Compares RFQ and Quote text using GPT-4o, enforces JSON output, 
#     and returns the result as a Python dictionary.
#     """
#     # Ensure API Key is set
#     openai.api_key = os.getenv("OPENAI_API_KEY")
#     if not openai.api_key:
#         raise EnvironmentError("OPENAI_API_KEY not found in environment variables.")

#     # Define the desired JSON structure clearly in the prompt and system message
#     json_structure = {
#         "items_requested": "[list of items requested in RFQ]",
#         "items_quoted": "[list of items quoted]",
#         "differences_in_items": "[list of differences in items, including alternates or EOL items]",
#         "discrepancies": "[list of pricing/term discrepancies]",
#         "potential_issues": "[list of potential issues, e.g., EOL parts]",
#         "summary": "summary of differences"
#     }
    
#     # Use f-string for the prompt for readability
#     prompt_content = f"""
#     You are an expert in analyzing RFQ and Quote documents.
#     Given the following RFQ and Quote texts, identify discrepancies in pricing, terms, and item specifications. 
#     - Make sure the quoted items are the same as the requested items in the RFQ.
#     - If items are different, highlight the differences.
#     - If the items are alternatives, mention that.
#     - If end-of-life (EOL) items are quoted, mention that.
#     - If the RFQ contained an EOL item and an alternative was quoted, mention that.

#     Provide the results **strictly** in the following JSON format: {json.dumps(json_structure, indent=2)}

#     RFQ Text:
#     ---
#     {rfq_text}
#     ---
#     Quote Text:
#     ---
#     {quote_text}
#     ---
#     """
    
#     try:
#         response = openai.ChatCompletion.create(
#             model="gpt-4o",
#             messages=[
#                 {"role": "system", 
#                  "content": "You are a specialized JSON output assistant. Your only task is to analyze the provided texts and return the analysis strictly as a single JSON object. DO NOT include any text outside the JSON object."},
#                 {"role": "user", "content": prompt_content}
#             ],
#             temperature=0,
#             max_tokens=1500,
#             # CRITICAL: Enforce JSON response format
#             response_format={"type": "json_object"}
#         )
        
#         # The response message content is now guaranteed (or highly likely) to be a JSON string
#         json_string = response.choices[0].message.content
        
#         # Parse the JSON string into a Python dictionary
#         result_dict = json.loads(json_string)
        
#         return result_dict
        
#     except openai.error.OpenAIError as e:
#         # Handle API errors gracefully
#         print(f"OpenAI API Error: {e}")
#         return {"error": "OpenAI API call failed", "details": str(e)}
#     except json.JSONDecodeError as e:
#         # Handle cases where the model did not return valid JSON
#         print(f"JSON Decoding Error: {e}")
#         print(f"Raw response: {json_string}")
#         return {"error": "Failed to parse JSON response from AI model", "details": str(e)}
#     except Exception as e:
#         # Handle other unexpected errors
#         print(f"An unexpected error occurred: {e}")
#         return {"error": "An internal error occurred", "details": str(e)}
    
    
# from werkzeug.datastructures import FileStorage

# def get_file_extension(filename):
#     """Returns the file extension in lowercase."""
#     return filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

# def parse_file_content(file: FileStorage):
#     """
#     Reads the file stream and parses content into a readable/structured format.
#     Returns the parsed data (e.g., DataFrame or String) or raises an error.
#     """
#     ext = get_file_extension(file.filename)
#     file_bytes = file.stream.read()

#     if ext in ['xlsx', 'xls', 'csv']:
#         # Handle Excel/CSV files using Pandas
#         file_stream = io.BytesIO(file_bytes)
        
#         if ext == 'csv':
#             # Assuming standard CSV encoding
#             return pd.read_csv(file_stream)
#         else:
#             # Assumes Excel file
#             return pd.read_excel(file_stream)
    
#     elif ext == 'pdf':
#         # --- PDF Parsing Placeholder ---
#         # NOTE: PDF parsing is complex and requires an external library.
#         # This is a placeholder; you must implement the PDF text extraction here.
        
#         # Example using pypdf:
#         # reader = PdfReader(io.BytesIO(file_bytes))
#         # text = ""
#         # for page in reader.pages:
#         #     text += page.extract_text()
#         # return text 
        
#         # For demonstration, we'll return a simple message if the file is PDF
#         return f"PDF data (requires parsing): {file.filename}" 

#     elif ext in ['txt']:
#         # Simple text file
#         return file_bytes.decode('utf-8')
        
#     else:
#         raise ValueError(f"Unsupported file type: .{ext}")



